import nltk
import os
from openpyxl import load_workbook


def clear_column():
    for i in range(2, ws.max_row+1):
        ws['M' + str(i)] = None
    wb.save(file_path)
    quit()


count = 2
num_1 = 0
num_students = 0

file_path = input("Enter excel file name: ") + ".xlsx"
if not os.path.isfile(file_path):
    print('File does not exist! Quitting program.')
    quit()
wb = load_workbook(file_path, data_only=True)
ws = wb.active

with open("names.txt", "r") as names_file:
    if (input("Clear column? Y/N: ").upper() == "Y"):
        clear_column()
    # For each name in the file
    for name in names_file:
        name = name.upper()
        # For each row in spreadsheet
        for i in range(2, ws.max_row+1):
            # Full name taken from column 1 and 2
            parsed_name = ws.cell(row=i, column=2).value.upper(
            ) + " " + ws.cell(row=i, column=1).value.upper()
            # If name from file is only first name or full name's edit distance is at most 1, then set cell value to 1
            # " " not in name and ws.cell(row=i, column=2).value.upper())
            if nltk.edit_distance(parsed_name, name) < 2:
                ws['M' + str(i)] = 1
        num_students += 1
    # Add 0's for each empty remaining cell
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=13).value is None:
            ws['M' + str(i)] = 0

wb.save(file_path)
print("\nWriting completed!")
print("Number of students who completed discussion board: " + str(num_students))
